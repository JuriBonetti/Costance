#_*_ coding: utf-8 _*_
import os
import pandas as pd
from openpyxl import load_workbook

print("Cartella corrente:", os.getcwd())

# Percorsi dei file
file_dati = "Dati_Ingresso.xlsx"  # File con i dati originali
file_risultati = "File_DaPopolare.xlsx"  # File dove scrivere il risultato

# Carica i dati
df = pd.read_excel(file_dati)

# Converte la colonna 'data' in formato datetime
df['data'] = pd.to_datetime(df['data'], dayfirst = True)

# Filtra solo il mese desiderato (es. Febbraio 2025)
mese_target = "2025-02"  # Cambia qui per un altro mese
df['mese'] = df['data'].dt.to_period('M')
df_filtro = df[(df['componente'] == 'Fosforo') & (df['mese'] == mese_target)]

# Calcola la media per il mese scelto
media_fosforo_valore = df_filtro['quantita'].mean() if not df_filtro.empty else "Nessun dato"

# Apri il file dei risultati senza alterare altro
wb = load_workbook(file_risultati)
ws = wb.active  # Usa il foglio attivo (cambia se serve un altro foglio)

# Scrivi la media nella cella B3
ws["B3"] = media_fosforo_valore

# Salva il file
wb.save(file_risultati)
wb.close()

print(df_filtro)
print(f"Media del Fosforo per {mese_target} scritta in 'File_DaPopolare.xlsx' (B3).")

input ("Premi invio per continuare...")

try:
    print("Media del fosforo scritta in 'File_DaPopolare.xlsx' (B3).")
    input("Premi INVIO per uscire...")
except Exception as e:
    print(f"Errore: {e}")
    input("Errore rilevato. Premi INVIO per chiudere...")