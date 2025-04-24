import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from shutil import copyfile
import calendar
from datetime import datetime
import os

# Mappa dei mesi alle colonne Excel
mesi_colonne = {
    1: 'B', 2: 'C', 3: 'D', 4: 'E', 5: 'F', 6: 'G',
    7: 'H', 8: 'I', 9: 'J', 10: 'K', 11: 'L', 12: 'M'
}

# Mappa dei parametri alle righe nei due file
parametri_info = {
    "Portata": {"file": "ingresso", "riga": 6},
    "Domanda Biochimica Ossigeno": {"ingresso": 11, "uscita": 12},
    "Domanda Chimica Ossigeno": {"ingresso": 7, "uscita": 8},
    "Azoto Totale": {"ingresso": 15, "uscita": 16},
    "Fosforo Totale": {"ingresso": 19, "uscita": 20}
}

# Carica dati da Excel
def carica_excel(file):
    try:
        df = pd.read_excel(file)
        df = df.rename(columns={
            'Nome parametro': 'componente',
            'Data Prelievo': 'data',
            'Risultato numerico': 'quantita'
        })
        if not all(col in df.columns for col in ['componente', 'data', 'quantita']):
            st.warning("Il file caricato non contiene le intestazioni richieste ('Nome parametro', 'Data Prelievo', 'Risultato numerico').")
            return pd.DataFrame()
        df['componente'] = df['componente'].str.title()
        df['data'] = pd.to_datetime(df['data'], dayfirst=True)
        return df
    except Exception as e:
        st.error(f"Errore caricando il file: {e}")
        return pd.DataFrame()

# Calcola media di un parametro per un mese
def calcola_media(df, parametro, mese):
    parametro = parametro.title()
    if 'data' not in df.columns or 'componente' not in df.columns or 'quantita' not in df.columns:
        st.warning(f"Colonne mancanti nei dati per il parametro '{parametro}'.")
        return None
    mese_period = pd.to_datetime(mese).to_period('M')
    df['mese'] = df['data'].dt.to_period('M')
    df_filtrato = df[df['mese'] == mese_period]
    df_param = df_filtrato[df_filtrato['componente'] == parametro]
    if df_param.empty:
        return None
    return df_param['quantita'].mean()

# Scrivi una media in una cella specifica
def scrivi_media(wb, colonna, riga, valore):
    ws = wb.active
    ws[f"{colonna}{riga}"] = valore

# Streamlit UI
st.title("Calcolo Medie Parametri Mensili")

# Upload dei file di ingresso e uscita
file_ingresso = st.file_uploader("Carica il file di ingresso", type="xlsx", key="ingresso")
file_uscita = st.file_uploader("Carica il file di uscita", type="xlsx", key="uscita")

# File fisso come modello di output
file_modello = "KPI_Ravenna_Master.xlsx"

# Inizializza la tabella dati in session_state
if "tabella_parametri" not in st.session_state:
    st.session_state.tabella_parametri = pd.DataFrame(columns=["Parametro", "Origine", "Mese"])

# Caricamento e selezione dinamica dei parametri
parametri_possibili = []
if file_ingresso is not None:
    df_ingresso_temp = carica_excel(file_ingresso)
    parametri_possibili += df_ingresso_temp['componente'].unique().tolist()
if file_uscita is not None:
    df_uscita_temp = carica_excel(file_uscita)
    parametri_possibili += df_uscita_temp['componente'].unique().tolist()
parametri_possibili = list(set(parametri_possibili))

# Aggiunta parametri alla tabella
col1, col2, col3 = st.columns(3)
with col1:
    parametro_input = st.selectbox("Parametro", parametri_possibili)
with col2:
    origine_input = st.selectbox("Origine del dato", ["ingresso", "uscita"])
with col3:
    mese_input_mese = st.selectbox("Mese", list(calendar.month_name)[1:])
    mese_input_anno = st.number_input("Anno", min_value=2000, max_value=2100, value=datetime.today().year)
    mese_input = pd.to_datetime(f"01 {mese_input_mese} {mese_input_anno}", dayfirst=True)

if st.button("Aggiungi alla tabella"):
    nuova_riga = pd.DataFrame([[parametro_input.title(), origine_input, mese_input]], columns=["Parametro", "Origine", "Mese"])
    st.session_state.tabella_parametri = pd.concat([st.session_state.tabella_parametri, nuova_riga], ignore_index=True)

# Mostra tabella dei parametri
st.subheader("Parametri selezionati")
st.dataframe(st.session_state.tabella_parametri)

# Calcolo e scrittura medie
if st.button("Calcola e Scrivi Medie"):
    if file_ingresso is None or file_uscita is None:
        st.error("Carica entrambi i file di ingresso e di uscita.")
    elif st.session_state.tabella_parametri.empty:
        st.error("Aggiungi almeno un parametro alla tabella.")
    else:
        df_ingresso = carica_excel(file_ingresso)
        df_uscita = carica_excel(file_uscita)

        # Crea una copia del file fisso KPI_Ravenna_Master.xlsx
        file_copia = "KPI_Ravenna_Master_modificato.xlsx"
        try:
            copyfile(file_modello, file_copia)
        except FileNotFoundError:
            st.error(f"Il file modello fisso '{file_modello}' non e stato trovato nella cartella.")
            st.stop()

        wb = load_workbook(file_copia)

        for _, row in st.session_state.tabella_parametri.iterrows():
            parametro = row["Parametro"].title()
            origine = row["Origine"]
            mese = row["Mese"]
            col_mese = mesi_colonne[pd.to_datetime(mese).month]

            df = df_ingresso if origine == "ingresso" else df_uscita
            if df is None:
                st.warning(f"File {origine} non disponibile per il parametro {parametro}.")
                continue

            media = calcola_media(df, parametro, mese)

            if media is not None:
                try:
                    riga = parametri_info[parametro][origine] if parametro != "Portata" else parametri_info["Portata"]["riga"]
                    scrivi_media(wb, col_mese, riga, media)
                    st.success(f"Scritta media {parametro} ({origine}) in {col_mese}{riga}")
                except KeyError:
                    st.warning(f"Il parametro '{parametro}' non ha una riga definita nella mappa parametri_info.")
            else:
                st.warning(f"Nessun dato per {parametro} ({origine}) nel mese selezionato")

        wb.save(file_copia)
        wb.close()

        st.info(f"File aggiornato salvato come: {file_copia}")
